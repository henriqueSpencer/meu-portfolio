import re
import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from io import BytesIO

from ..database import get_db
from ..core.security import get_current_user
from ..models.user import User
from ..models.transaction import Transaction
from ..models.br_stock import BrStock
from ..models.fii import Fii
from ..models.fi_etf import FiEtf
from ..schemas.import_b3 import (
    ImportedRow,
    ImportPreviewResponse,
    ImportSummary,
    ImportConfirmRequest,
    ImportConfirmResponse,
)
from ..services.yahoo import fetch_asset_info
from .transactions import _apply

router = APIRouter(prefix="/api/import/b3", tags=["import-b3"])

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FI_ETF_TICKERS = {
    "LFTB11", "LFTS11", "NTNS11", "NTNB11", "IMAB11", "IRFM11",
    "B5P211", "IB5M11", "FIXA11", "KDIF11", "IDKA11",
}

# Stock units (tickers ending in 11 that are stocks, not FIIs)
STOCK_UNITS = {
    "BIDI11", "TAEE11", "KLBN11", "SANB11", "SAPR11", "ENBR11",
    "BPAC11", "ENGI11", "EGIE11", "ALUP11", "UNIT11", "RNEW11",
    "AESB11", "TIMS11",
}

FII_PATTERN = re.compile(r"^[A-Z]{4}1[1-3]B?$")

OPTION_MARKETS = {"Opção de Compra sobre Ações", "Opção de Venda sobre Ações"}

BROKER_MAP = {
    "BTG PACTUAL": "BTG Pactual",
    "XP INVESTIMENTOS": "XP",
    "CLEAR CORRETORA": "Clear",
    "RICO INVESTIMENTOS": "Rico",
    "INTER DISTRIBUIDORA": "Inter",
    "NU INVEST": "Nu Invest",
    "MODAL": "Modal",
    "GENIAL": "Genial",
    "ATIVA INVESTIMENTOS": "Ativa",
    "GUIDE INVESTIMENTOS": "Guide",
    "EASYNVEST": "Easynvest",
    "ORAMA": "Orama",
    "TERRA INVESTIMENTOS": "Terra",
    "MIRAE ASSET": "Mirae",
    "TORO INVESTIMENTOS": "Toro",
    "WARREN": "Warren",
    "BANCO DO BRASIL": "BB",
    "ITAU": "Itau",
    "BRADESCO": "Bradesco",
    "SANTANDER": "Santander",
    "CAIXA": "Caixa",
    "SAFRA": "Safra",
    "AGORA": "Agora",
}


def _abbreviate_broker(raw: str) -> str:
    upper = raw.upper()
    for key, short in BROKER_MAP.items():
        if key in upper:
            return short
    # Fallback: first two words
    parts = raw.split()
    return " ".join(parts[:2]) if len(parts) >= 2 else raw


def _classify_ticker(ticker: str, market: str) -> tuple[str, str]:
    """Return (asset_class, clean_ticker)."""
    clean = ticker.upper().strip()

    # Mercado Fracionario: strip trailing F
    if market == "Mercado Fracionário" and clean.endswith("F") and len(clean) > 4:
        clean = clean[:-1]

    # FI ETFs (check first - some match FII pattern)
    if clean in FI_ETF_TICKERS:
        return "fi_etf", clean

    # FIIs: pattern XXXX1[1-3] (excluding known stock units)
    if FII_PATTERN.match(clean) and clean not in STOCK_UNITS:
        return "fii", clean

    # Everything else is br_stock
    return "br_stock", clean


def _parse_date(val) -> datetime.date:
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    # String format DD/MM/YYYY
    s = str(val).strip()
    parts = s.split("/")
    if len(parts) == 3:
        return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
    raise ValueError(f"Cannot parse date: {val}")


# ---------------------------------------------------------------------------
# Preview endpoint
# ---------------------------------------------------------------------------

@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_b3(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser .xlsx")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    rows_data: list[ImportedRow] = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue

        # Skip empty rows
        if not row[0]:
            continue

        date_val, tipo, mercado, _prazo, instituicao, codigo, qtd, preco, valor = row

        try:
            date = _parse_date(date_val)
        except ValueError:
            continue

        market = str(mercado or "").strip()
        ticker_raw = str(codigo or "").strip().upper()
        broker = _abbreviate_broker(str(instituicao or ""))

        # Skip options
        if market in OPTION_MARKETS:
            rows_data.append(ImportedRow(
                date=date,
                operation_type=str(tipo).lower().strip(),
                market=market,
                asset_class="br_stock",
                ticker=ticker_raw,
                qty=int(qtd or 0),
                unit_price=float(preco or 0),
                total_value=float(valor or 0),
                broker=broker,
                asset_name=ticker_raw,
                is_skipped=True,
                skip_reason=f"Opcao: {market}",
            ))
            continue

        # Classify ticker
        asset_class, clean_ticker = _classify_ticker(ticker_raw, market)
        operation = "compra" if str(tipo).strip().lower() == "compra" else "venda"

        rows_data.append(ImportedRow(
            date=date,
            operation_type=operation,
            market=market,
            asset_class=asset_class,
            ticker=clean_ticker,
            qty=int(qtd or 0),
            unit_price=round(float(preco or 0), 2),
            total_value=round(float(valor or 0), 2),
            broker=broker,
            asset_name=clean_ticker,
        ))

    wb.close()

    # --- Check duplicates against existing transactions ---
    result = await db.execute(select(Transaction).where(Transaction.user_id == user.id))
    existing_txs = result.scalars().all()
    existing_keys = set()
    for tx in existing_txs:
        key = (
            tx.date,
            (tx.ticker or "").upper(),
            tx.operation_type,
            int(tx.qty or 0),
            round(tx.unit_price or 0, 2),
        )
        existing_keys.add(key)

    # --- Check which assets exist ---
    br_result = await db.execute(select(BrStock.ticker).where(BrStock.user_id == user.id))
    br_tickers = {r[0] for r in br_result.all()}

    fii_result = await db.execute(select(Fii.ticker).where(Fii.user_id == user.id))
    fii_tickers = {r[0] for r in fii_result.all()}

    etf_result = await db.execute(select(FiEtf.ticker).where(FiEtf.user_id == user.id))
    etf_tickers = {r[0] for r in etf_result.all()}

    asset_sets = {
        "br_stock": br_tickers,
        "fii": fii_tickers,
        "fi_etf": etf_tickers,
    }

    new_assets = set()

    for row in rows_data:
        if row.is_skipped:
            continue

        # Check duplicate
        key = (row.date, row.ticker, row.operation_type, row.qty, row.unit_price)
        if key in existing_keys:
            row.is_duplicate = True

        # Check asset existence
        ticker_set = asset_sets.get(row.asset_class, set())
        if row.ticker in ticker_set:
            row.asset_exists = True
        else:
            new_assets.add(row.ticker)

    # Build summary
    active_rows = [r for r in rows_data if not r.is_skipped]
    summary = ImportSummary(
        total=len(rows_data),
        new=sum(1 for r in active_rows if not r.is_duplicate),
        duplicates=sum(1 for r in active_rows if r.is_duplicate),
        skipped=sum(1 for r in rows_data if r.is_skipped),
        new_assets=sorted(new_assets),
    )

    return ImportPreviewResponse(rows=rows_data, summary=summary)


# ---------------------------------------------------------------------------
# Confirm endpoint
# ---------------------------------------------------------------------------

@router.post("/confirm", response_model=ImportConfirmResponse)
async def confirm_b3(
    body: ImportConfirmRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Sort by date ASC for correct avg_price calculation
    sorted_rows = sorted(body.rows, key=lambda r: r.date)

    created_count = 0
    assets_created = []
    errors = []

    # Pre-load existing assets
    br_result = await db.execute(select(BrStock.ticker).where(BrStock.user_id == user.id))
    br_tickers = {r[0] for r in br_result.all()}
    fii_result = await db.execute(select(Fii.ticker).where(Fii.user_id == user.id))
    fii_tickers = {r[0] for r in fii_result.all()}
    etf_result = await db.execute(select(FiEtf.ticker).where(FiEtf.user_id == user.id))
    etf_tickers = {r[0] for r in etf_result.all()}

    asset_sets = {
        "br_stock": br_tickers,
        "fii": fii_tickers,
        "fi_etf": etf_tickers,
    }

    # Pre-fetch sector info from Yahoo Finance for new tickers
    new_tickers_to_lookup = set()
    for row in sorted_rows:
        ticker_set = asset_sets.get(row.asset_class, set())
        if row.ticker not in ticker_set and row.asset_class in ("br_stock", "fii"):
            new_tickers_to_lookup.add(row.ticker)

    asset_info_map = {}
    if new_tickers_to_lookup:
        try:
            asset_info_map = await fetch_asset_info(list(new_tickers_to_lookup))
        except Exception:
            pass  # Fallback: will use "A classificar"

    for row in sorted_rows:
        try:
            # Auto-create asset if missing
            ticker_set = asset_sets.get(row.asset_class, set())
            if row.ticker not in ticker_set:
                info = asset_info_map.get(row.ticker, {})
                sector = info.get("sector") or "A classificar"
                name = info.get("name") or row.ticker

                if row.asset_class == "br_stock":
                    asset = BrStock(
                        ticker=row.ticker,
                        name=name,
                        sector=sector,
                        qty=0,
                        avg_price=0,
                        current_price=0,
                        broker=row.broker,
                        user_id=user.id,
                    )
                    db.add(asset)
                elif row.asset_class == "fii":
                    asset = Fii(
                        ticker=row.ticker,
                        name=name,
                        sector=sector,
                        qty=0,
                        avg_price=0,
                        current_price=0,
                        broker=row.broker,
                        user_id=user.id,
                    )
                    db.add(asset)
                elif row.asset_class == "fi_etf":
                    asset = FiEtf(
                        ticker=row.ticker,
                        name=name,
                        qty=0,
                        avg_price=0,
                        current_price=0,
                        broker=row.broker,
                        user_id=user.id,
                    )
                    db.add(asset)

                ticker_set.add(row.ticker)
                assets_created.append(row.ticker)
                await db.flush()

            # Create transaction
            tx = Transaction(
                date=row.date,
                operation_type=row.operation_type,
                asset_class=row.asset_class,
                ticker=row.ticker,
                asset_name=row.asset_name,
                qty=row.qty,
                unit_price=row.unit_price,
                total_value=row.total_value,
                broker=row.broker,
                fees=0,
                notes=f"Importado B3 - {row.market}",
                user_id=user.id,
            )
            db.add(tx)
            await db.flush()

            # Apply position changes
            await _apply(db, tx, user.id)
            created_count += 1

        except Exception as e:
            errors.append(f"{row.ticker} ({row.date}): {str(e)}")

    await db.commit()

    return ImportConfirmResponse(
        created=created_count,
        assets_created=sorted(set(assets_created)),
        errors=errors,
    )
