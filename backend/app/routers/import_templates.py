from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from ..core.security import get_current_user
from ..models.user import User

router = APIRouter(prefix="/api/import/templates", tags=["import-templates"])

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
EXAMPLE_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _style_example_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


@router.get("/negociacao")
async def template_negociacao(user: User = Depends(get_current_user)):
    """Template for B3 negotiation statement (negociacao)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Negociacao"

    headers = [
        "Data",
        "Tipo",
        "Mercado",
        "Prazo",
        "Instituicao",
        "Codigo",
        "Quantidade",
        "Preco",
        "Valor",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    examples = [
        ["15/01/2026", "Compra", "Mercado a Vista", "", "XP INVESTIMENTOS", "PETR4", 100, 38.50, 3850.00],
        ["15/01/2026", "Venda", "Mercado Fracionario", "", "BTG PACTUAL", "VALE3", 50, 62.30, 3115.00],
        ["20/01/2026", "Compra", "Mercado a Vista", "", "INTER DISTRIBUIDORA", "HGLG11", 10, 165.00, 1650.00],
        ["20/01/2026", "Compra", "Mercado a Vista", "", "NU INVEST", "LFTS11", 5, 110.00, 550.00],
    ]
    for i, row in enumerate(examples):
        ws.append(row)
        _style_example_row(ws, i + 2, len(headers))

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-negociacao-b3.xlsx"},
    )


@router.get("/movimentacao")
async def template_movimentacao(user: User = Depends(get_current_user)):
    """Template for B3 movimentacao statement."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentacao"

    headers = [
        "Entrada/Saida",
        "Data",
        "Movimentacao",
        "Produto",
        "Instituicao",
        "Quantidade",
        "Preco unitario",
        "Valor da Operacao",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    examples = [
        ["Credito", "15/01/2026", "Dividendo", "PETR4 - PETROBRAS PN", "XP INVESTIMENTOS", 100, 0, 150.00],
        ["Credito", "15/01/2026", "Juros Sobre Capital Proprio", "ITUB4 - ITAU UNIBANCO PN", "BTG PACTUAL", 200, 0, 80.00],
        ["Credito", "20/01/2026", "Rendimento", "HGLG11 - CSHG LOGISTICA FII", "INTER DISTRIBUIDORA", 10, 0, 8.50],
        ["Credito", "10/01/2026", "Transferencia - Liquidacao", "CDB - CDB1234X56 - BANCO XYZ", "NU INVEST", 1, 1000.00, 1000.00],
        ["Credito", "05/01/2026", "Vencimento", "CDB - CDB9876Y12 - BANCO ABC", "XP INVESTIMENTOS", 1, 0, 5250.00],
        ["Credito", "01/02/2026", "Bonificacao em Ativos", "VALE3 - VALE ON", "BTG PACTUAL", 5, 0, 0],
    ]
    for i, row in enumerate(examples):
        ws.append(row)
        _style_example_row(ws, i + 2, len(headers))

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=movimentacao-template-b3.xlsx"},
    )


@router.get("/backup")
async def template_backup(user: User = Depends(get_current_user)):
    """Template for backup import/restore."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Backup"

    headers = [
        "date",
        "operation_type",
        "asset_class",
        "ticker",
        "asset_id",
        "asset_name",
        "qty",
        "unit_price",
        "total_value",
        "broker",
        "broker_destination",
        "fees",
        "notes",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    examples = [
        ["2026-01-15", "compra", "br_stock", "PETR4", "", "Petrobras PN", 100, 38.50, 3850.00, "XP", "", 0, ""],
        ["2026-01-15", "compra", "fii", "HGLG11", "", "CSHG Logistica FII", 10, 165.00, 1650.00, "BTG Pactual", "", 0, ""],
        ["2026-01-20", "aporte", "fixed_income", "", "uuid-do-ativo", "CDB Banco XYZ 120% CDI", 1, 0, 10000.00, "Nu Invest", "", 0, "Venc. 2028"],
        ["2026-01-20", "compra", "fi_etf", "LFTS11", "", "LFTS11", 5, 110.00, 550.00, "Inter", "", 0, ""],
        ["2026-02-01", "venda", "br_stock", "VALE3", "", "Vale ON", 50, 62.30, 3115.00, "BTG Pactual", "", 4.90, "Venda parcial"],
        ["2026-02-05", "transferencia", "br_stock", "PETR4", "", "Petrobras PN", 100, 38.50, 3850.00, "XP", "BTG Pactual", 0, ""],
    ]
    for i, row in enumerate(examples):
        ws.append(row)
        _style_example_row(ws, i + 2, len(headers))

    _auto_width(ws)

    # Add instructions sheet
    ws2 = wb.create_sheet("Instrucoes")
    instructions = [
        ["Campo", "Descricao", "Valores aceitos"],
        ["date", "Data da operacao", "YYYY-MM-DD ou DD/MM/YYYY"],
        ["operation_type", "Tipo de operacao", "compra, venda, aporte, resgate, transferencia, desdobramento, bonificacao"],
        ["asset_class", "Classe do ativo", "br_stock, fii, intl_stock, fixed_income, fi_etf, cash_account, real_asset"],
        ["ticker", "Ticker (acoes, FIIs, ETFs)", "Ex: PETR4, HGLG11, LFTS11, AAPL"],
        ["asset_id", "ID do ativo (RF, caixa, imobilizado)", "UUID do ativo no sistema (deixe vazio para ativos com ticker)"],
        ["asset_name", "Nome do ativo", "Nome legivel do ativo"],
        ["qty", "Quantidade", "Numero (pode ser decimal)"],
        ["unit_price", "Preco unitario", "Valor em R$"],
        ["total_value", "Valor total", "Valor em R$"],
        ["broker", "Corretora de origem", "Nome da corretora"],
        ["broker_destination", "Corretora destino (so transferencia)", "Nome da corretora destino"],
        ["fees", "Taxas/custos", "Valor em R$"],
        ["notes", "Observacoes", "Texto livre"],
    ]
    for r, row in enumerate(instructions):
        ws2.append(row)
        if r == 0:
            for c in range(1, 4):
                cell = ws2.cell(row=1, column=c)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
        else:
            for c in range(1, 4):
                ws2.cell(row=r + 1, column=c).border = THIN_BORDER

    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-backup.xlsx"},
    )
