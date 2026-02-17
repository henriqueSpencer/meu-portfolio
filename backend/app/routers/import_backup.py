import datetime
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from io import BytesIO
from pydantic import BaseModel

from ..database import get_db
from ..models.transaction import Transaction
from ..models.br_stock import BrStock
from ..models.fii import Fii
from ..models.intl_stock import IntlStock
from ..models.fixed_income import FixedIncome
from ..models.fi_etf import FiEtf
from ..models.cash_account import CashAccount
from ..models.real_asset import RealAsset
from ..services.yahoo import fetch_asset_info
from .transactions import _apply

router = APIRouter(prefix="/api/import/backup", tags=["import-backup"])

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class BackupRow(BaseModel):
    date: str
    operation_type: str
    asset_class: str
    ticker: str | None = None
    asset_id: str | None = None
    asset_name: str
    qty: float | None = None
    unit_price: float | None = None
    total_value: float | None = None
    broker: str = ""
    broker_destination: str | None = None
    fees: float = 0
    notes: str | None = None
    is_duplicate: bool = False


class BackupPreviewResponse(BaseModel):
    rows: list[BackupRow]
    summary: dict


class BackupConfirmRequest(BaseModel):
    rows: list[BackupRow]


class BackupConfirmResponse(BaseModel):
    created: int
    assets_created: list[str]
    errors: list[str]


# Column order matching the export
COL_MAP = [
    "date", "operation_type", "asset_class", "ticker", "asset_id",
    "asset_name", "qty", "unit_price", "total_value", "broker",
    "broker_destination", "fees", "notes",
]


def _parse_val(val, col_name: str):
    """Parse a cell value based on column name."""
    if val is None:
        return None

    if col_name == "date":
        if isinstance(val, datetime.datetime):
            return val.date().isoformat()
        if isinstance(val, datetime.date):
            return val.isoformat()
        s = str(val).strip()
        # Try ISO format (YYYY-MM-DD)
        if len(s) == 10 and s[4] == "-":
            return s
        # Try DD/MM/YYYY
        parts = s.split("/")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return s

    if col_name in ("qty", "unit_price", "total_value", "fees"):
        try:
            return float(val) if val else None
        except (ValueError, TypeError):
            return None

    return str(val).strip() if val else None


# ---------------------------------------------------------------------------
# Preview
# ---------------------------------------------------------------------------

@router.post("/preview", response_model=BackupPreviewResponse)
async def preview_backup(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser .xlsx")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    rows_data: list[BackupRow] = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue

        if not row or not row[0]:
            continue

        parsed = {}
        for col_idx, col_name in enumerate(COL_MAP):
            val = row[col_idx] if col_idx < len(row) else None
            parsed[col_name] = _parse_val(val, col_name)

        # Ensure required fields
        if not parsed.get("date") or not parsed.get("operation_type") or not parsed.get("asset_class"):
            continue

        # Apply defaults for fields that can't be None
        if not parsed.get("broker"):
            parsed["broker"] = ""
        if parsed.get("fees") is None:
            parsed["fees"] = 0
        if not parsed.get("asset_name"):
            parsed["asset_name"] = parsed.get("ticker") or parsed.get("asset_id") or ""

        rows_data.append(BackupRow(**parsed))

    wb.close()

    # Check duplicates against existing transactions
    result = await db.execute(select(Transaction))
    existing_txs = result.scalars().all()
    existing_keys = set()
    for tx in existing_txs:
        key = (
            tx.date.isoformat() if tx.date else "",
            (tx.ticker or tx.asset_id or ""),
            tx.operation_type,
            round(tx.qty or 0, 4),
            round(tx.unit_price or 0, 2),
        )
        existing_keys.add(key)

    duplicates = 0
    for row in rows_data:
        key = (
            row.date,
            (row.ticker or row.asset_id or ""),
            row.operation_type,
            round(row.qty or 0, 4),
            round(row.unit_price or 0, 2),
        )
        if key in existing_keys:
            row.is_duplicate = True
            duplicates += 1

    summary = {
        "total": len(rows_data),
        "new": len(rows_data) - duplicates,
        "duplicates": duplicates,
    }

    return BackupPreviewResponse(rows=rows_data, summary=summary)


# ---------------------------------------------------------------------------
# Confirm
# ---------------------------------------------------------------------------

# Maps asset_class -> (Model, pk_field, create_fn)
TICKER_CLASSES = {"br_stock", "fii", "intl_stock", "fi_etf"}
ID_CLASSES = {"fixed_income", "cash_account", "real_asset"}


def _make_asset(asset_class: str, row: BackupRow, asset_info_map: dict = None):
    """Create a minimal asset record from transaction data."""
    info = (asset_info_map or {}).get(row.ticker, {})

    if asset_class == "br_stock":
        return BrStock(
            ticker=row.ticker,
            name=info.get("name") or row.asset_name or row.ticker,
            sector=info.get("sector") or "A classificar",
            qty=0, avg_price=0, current_price=0,
            broker=row.broker,
        )
    elif asset_class == "fii":
        return Fii(
            ticker=row.ticker,
            name=info.get("name") or row.asset_name or row.ticker,
            sector=info.get("sector") or "A classificar",
            qty=0, avg_price=0, current_price=0,
            broker=row.broker,
        )
    elif asset_class == "fi_etf":
        return FiEtf(
            ticker=row.ticker,
            name=info.get("name") or row.asset_name or row.ticker,
            qty=0, avg_price=0, current_price=0,
            broker=row.broker,
        )
    elif asset_class == "intl_stock":
        return IntlStock(
            ticker=row.ticker,
            name=info.get("name") or row.asset_name or row.ticker,
            sector=info.get("sector") or "A classificar",
            qty=0, avg_price_usd=0, current_price_usd=0,
            broker=row.broker,
        )
    elif asset_class == "fixed_income":
        new_id = row.asset_id or str(uuid.uuid4())
        return FixedIncome(
            id=new_id,
            title=row.asset_name or "Importado",
            type="CDB",
            rate="CDI 100%",
            applied_value=0,
            current_value=0,
            application_date=datetime.date.fromisoformat(row.date),
            maturity_date=datetime.date.fromisoformat(row.date) + datetime.timedelta(days=365),
            broker=row.broker,
        ), new_id
    elif asset_class == "cash_account":
        new_id = row.asset_id or str(uuid.uuid4())
        return CashAccount(
            id=new_id,
            name=row.asset_name or "Importado",
            type="conta_corrente",
            institution=row.broker,
            balance=0,
        ), new_id
    elif asset_class == "real_asset":
        new_id = row.asset_id or str(uuid.uuid4())
        return RealAsset(
            id=new_id,
            description=row.asset_name or "Importado",
            type="Imovel",
            estimated_value=0,
            acquisition_date=datetime.date.fromisoformat(row.date),
        ), new_id
    return None


@router.post("/confirm", response_model=BackupConfirmResponse)
async def confirm_backup(
    body: BackupConfirmRequest,
    db: AsyncSession = Depends(get_db),
):
    # Sort by date ASC for correct position calculation
    sorted_rows = sorted(body.rows, key=lambda r: r.date)

    created_count = 0
    assets_created = []
    errors = []

    # Pre-load existing assets (tickers and IDs)
    existing_assets: dict[str, set[str]] = {}

    for model, field in [
        (BrStock, "ticker"), (Fii, "ticker"), (IntlStock, "ticker"), (FiEtf, "ticker"),
    ]:
        result = await db.execute(select(getattr(model, field)))
        existing_assets[model.__tablename__] = {r[0] for r in result.all()}

    for model, field in [
        (FixedIncome, "id"), (CashAccount, "id"), (RealAsset, "id"),
    ]:
        result = await db.execute(select(getattr(model, field)))
        existing_assets[model.__tablename__] = {str(r[0]) for r in result.all()}

    TABLE_MAP = {
        "br_stock": "br_stocks",
        "fii": "fiis",
        "intl_stock": "intl_stocks",
        "fi_etf": "fi_etfs",
        "fixed_income": "fixed_income",
        "cash_account": "cash_accounts",
        "real_asset": "real_assets",
    }

    # Pre-fetch sector info from Yahoo Finance for new tickers
    new_br_tickers = set()
    new_intl_tickers = set()
    for row in sorted_rows:
        ac = row.asset_class
        table_name = TABLE_MAP.get(ac)
        if not table_name or not row.ticker:
            continue
        if ac in ("br_stock", "fii", "fi_etf") and row.ticker not in existing_assets.get(table_name, set()):
            new_br_tickers.add(row.ticker)
        elif ac == "intl_stock" and row.ticker not in existing_assets.get(table_name, set()):
            new_intl_tickers.add(row.ticker)

    asset_info_map = {}
    try:
        if new_br_tickers:
            asset_info_map.update(await fetch_asset_info(list(new_br_tickers), suffix=".SA"))
        if new_intl_tickers:
            asset_info_map.update(await fetch_asset_info(list(new_intl_tickers), suffix=""))
    except Exception:
        pass  # Fallback: will use "A classificar"

    # Track newly-assigned IDs so transactions reference the right asset
    id_remap: dict[str, str] = {}

    for row in sorted_rows:
        try:
            asset_class = row.asset_class
            table_name = TABLE_MAP.get(asset_class)
            if not table_name:
                errors.append(f"Classe desconhecida: {asset_class}")
                continue

            asset_key = row.ticker if asset_class in TICKER_CLASSES else (row.asset_id or "")

            # Auto-create asset if missing
            if asset_key and asset_key not in existing_assets.get(table_name, set()):
                result = _make_asset(asset_class, row, asset_info_map)
                if result is None:
                    errors.append(f"Nao foi possivel criar ativo: {asset_class}/{asset_key}")
                    continue

                if asset_class in ID_CLASSES:
                    asset_obj, new_id = result
                    if new_id != row.asset_id:
                        id_remap[row.asset_id or ""] = new_id
                    db.add(asset_obj)
                    existing_assets[table_name].add(new_id)
                    asset_key = new_id
                    assets_created.append(row.asset_name or new_id)
                else:
                    db.add(result)
                    existing_assets[table_name].add(asset_key)
                    assets_created.append(asset_key)

                await db.flush()

            # Resolve asset_id (may have been remapped)
            actual_asset_id = row.asset_id
            if asset_class in ID_CLASSES and row.asset_id in id_remap:
                actual_asset_id = id_remap[row.asset_id]

            # Create transaction
            tx = Transaction(
                date=datetime.date.fromisoformat(row.date),
                operation_type=row.operation_type,
                asset_class=asset_class,
                ticker=row.ticker,
                asset_id=actual_asset_id,
                asset_name=row.asset_name,
                qty=row.qty,
                unit_price=row.unit_price,
                total_value=row.total_value,
                broker=row.broker,
                broker_destination=row.broker_destination,
                fees=row.fees or 0,
                notes=row.notes,
            )
            db.add(tx)
            await db.flush()

            # Apply position changes
            await _apply(db, tx)
            created_count += 1

        except Exception as e:
            errors.append(f"{row.ticker or row.asset_name} ({row.date}): {str(e)}")

    await db.commit()

    return BackupConfirmResponse(
        created=created_count,
        assets_created=sorted(set(assets_created)),
        errors=errors,
    )
