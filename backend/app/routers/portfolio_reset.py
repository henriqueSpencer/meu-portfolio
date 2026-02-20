import io
import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import select, delete, func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..database import get_db
from ..models.user import User
from ..models.transaction import Transaction
from ..models.dividend import Dividend
from ..models.br_stock import BrStock
from ..models.fii import Fii
from ..models.intl_stock import IntlStock
from ..models.fixed_income import FixedIncome
from ..models.fi_etf import FiEtf
from ..models.cash_account import CashAccount
from ..models.real_asset import RealAsset
from ..models.patrimonial_history import PatrimonialHistory
from ..models.watchlist import WatchlistItem
from ..models.allocation_target import AllocationTarget
from ..models.accumulation_goal import AccumulationGoal
from ..core.security import get_current_user

router = APIRouter(prefix="/api/portfolio", tags=["portfolio"])

# Column definitions for the export XLSX
TX_COLUMNS = [
    "date", "operation_type", "asset_class", "ticker", "asset_id",
    "asset_name", "qty", "unit_price", "total_value", "broker",
    "broker_destination", "fees", "notes",
]

TX_HEADERS = [
    "Data", "Operacao", "Classe", "Ticker", "Asset ID",
    "Nome do Ativo", "Qtd", "Preco Unitario", "Valor Total", "Corretora",
    "Corretora Destino", "Taxas", "Notas",
]


@router.get("/export")
async def export_transactions(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Export all transactions as an XLSX file."""
    result = await db.execute(
        select(Transaction)
        .where(Transaction.user_id == user.id)
        .order_by(Transaction.date.asc(), Transaction.id.asc())
    )
    txs = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lancamentos"

    # --- Styling ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write header
    for col_idx, header in enumerate(TX_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, tx in enumerate(txs, 2):
        for col_idx, col_name in enumerate(TX_COLUMNS, 1):
            val = getattr(tx, col_name, None)
            if isinstance(val, datetime.date):
                val = val.isoformat()
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width columns
    for col_idx in range(1, len(TX_COLUMNS) + 1):
        max_len = len(TX_HEADERS[col_idx - 1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"backup_lancamentos_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# All tables to clear during a full portfolio reset (user-scoped)
ALL_MODELS = [
    Transaction, Dividend, PatrimonialHistory,
    WatchlistItem, AllocationTarget, AccumulationGoal,
    BrStock, Fii, IntlStock, FixedIncome, FiEtf, CashAccount, RealAsset,
]


@router.post("/reset")
async def reset_portfolio(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Delete ALL portfolio data for the current user."""
    counts = {}
    total = 0

    for model in ALL_MODELS:
        c = await db.scalar(
            select(func.count()).select_from(model).where(model.user_id == user.id)
        )
        await db.execute(delete(model).where(model.user_id == user.id))
        counts[model.__tablename__] = c
        total += c

    await db.commit()

    return {"deleted": counts, "total": total}
