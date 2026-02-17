import re
import datetime
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from io import BytesIO
from pydantic import BaseModel

from ..database import get_db
from ..models.transaction import Transaction
from ..models.dividend import Dividend
from ..models.fixed_income import FixedIncome
from ..models.br_stock import BrStock
from ..models.fii import Fii
from ..models.fi_etf import FiEtf
from ..services.yahoo import fetch_asset_info
from .transactions import _apply
from .import_b3 import _classify_ticker, _abbreviate_broker, _parse_date

router = APIRouter(prefix="/api/import/b3-mov", tags=["import-b3-mov"])

# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class MovRow(BaseModel):
    date: str  # YYYY-MM-DD
    direction: str  # Credito / Debito
    movement_type: str  # Original B3 movement type
    product: str  # Original product string
    institution: str
    qty: float | None = None
    unit_price: float | None = None
    total_value: float | None = None
    category: str  # provento, renda_fixa, evento, ignorado
    import_as: str  # dividendo, jcp, rendimento, compra_rf, vencimento_rf, resgate_rf, juros_rf, amortizacao_rf, bonificacao, desdobramento, venda, ignorado
    ticker: str | None = None
    asset_name: str = ""
    asset_class: str | None = None  # br_stock, fii, fi_etf, fixed_income
    rf_type: str | None = None  # CDB, CRA, CRI, DEB, Tesouro
    rf_code: str | None = None  # Specific bond code
    is_duplicate: bool = False
    is_skipped: bool = False
    skip_reason: str | None = None


class MovSummary(BaseModel):
    total: int
    proventos: int
    renda_fixa: int
    eventos: int
    ignorados: int
    duplicates: int
    new_assets: list[str]


class MovPreviewResponse(BaseModel):
    file_type: str = "movimentacao"
    rows: list[MovRow]
    summary: MovSummary


class MovConfirmRequest(BaseModel):
    rows: list[MovRow]


class MovConfirmResponse(BaseModel):
    dividends_created: int
    transactions_created: int
    assets_created: list[str]
    errors: list[str]


# ---------------------------------------------------------------------------
# Accent stripping
# ---------------------------------------------------------------------------


def _strip_accents(s: str) -> str:
    """Remove accents from a string: Bonificação -> Bonificacao."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


# ---------------------------------------------------------------------------
# Constants — movement type classification
# ---------------------------------------------------------------------------

SKIP_PATTERNS = [
    "cancelado", "transferencia", "transferencia - liquidacao",
    "emprestimo", "reembolso", "incorporacao", "atualizacao",
    "transferido",
]

SKIP_PREFIXES = [
    "cessao de direitos", "direito", "solicitacao", "recibo",
    "fracao", "leilao",
]

PROVENTO_MAP = {
    "dividendo": "dividendo",
    "juros sobre capital proprio": "jcp",
    "rendimento": "rendimento",
}

RF_COMPRA_TYPES = {
    "compra / venda",
    "compra/venda",
    "compra/venda definitiva/cessao",
    "aplicacao",
}

RF_TYPES_SET = {"CDB", "CRA", "CRI", "DEB", "LCA", "LCI", "LC"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _should_skip(movement_type: str, product: str) -> str | None:
    """Return skip reason or None."""
    mt_norm = _strip_accents(movement_type.lower().strip())
    prod_norm = _strip_accents(product.lower().strip())

    for pat in SKIP_PATTERNS:
        if pat in mt_norm:
            return f"Ignorado: {movement_type}"

    for prefix in SKIP_PREFIXES:
        if mt_norm.startswith(prefix):
            return f"Ignorado: {movement_type}"

    # Options
    if "opcao" in prod_norm or "opcao" in mt_norm:
        return f"Opcao: {product}"

    return None


def _parse_number(val) -> float | None:
    """Parse a number cell, returning None for '-' or empty."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("-", "", "0"):
        return None if s == "-" else (0.0 if s == "0" else None)
    # Handle Brazilian number format: 1.234,56
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _extract_ticker_from_product(product: str) -> dict:
    """Extract ticker/rf info from B3 product string.

    Examples:
      'ITSA4 - ITAUSA S/A'           -> ticker=ITSA4, asset_name=ITAUSA S/A
      'CDB - CDB8243X93D - QISTA...' -> rf_type=CDB, rf_code=CDB8243X93D
      'DEB - NTEN11 - NORTE ENERG...' -> rf_type=DEB, rf_code=NTEN11
      'Tesouro Selic 2025'            -> rf_type=Tesouro, asset_name=Tesouro Selic 2025
    """
    result = {
        "ticker": None,
        "asset_name": product,
        "asset_class": None,
        "rf_type": None,
        "rf_code": None,
    }

    product = product.strip()
    if not product:
        return result

    # Tesouro direto
    if product.lower().startswith("tesouro"):
        result["rf_type"] = "Tesouro"
        result["asset_class"] = "fixed_income"
        result["asset_name"] = product
        # Use a sanitized ID
        result["rf_code"] = re.sub(r"[^A-Za-z0-9]", "", product)[:36]
        return result

    parts = [p.strip() for p in product.split(" - ", 2)]

    # Fixed income: TYPE - CODE - NAME  or  TYPE - CODE
    if len(parts) >= 2 and parts[0].upper() in RF_TYPES_SET:
        result["rf_type"] = parts[0].upper()
        result["rf_code"] = parts[1].strip()[:36]
        result["asset_class"] = "fixed_income"
        result["asset_name"] = parts[2] if len(parts) >= 3 else parts[1]
        return result

    # Stock/FII: TICKER - NAME
    if len(parts) >= 2:
        ticker_candidate = parts[0].upper().strip()
        # Valid ticker pattern: 4-6 alphanumeric chars
        if re.match(r"^[A-Z]{3,5}\d{1,2}F?$", ticker_candidate):
            asset_class, clean_ticker = _classify_ticker(ticker_candidate, "")
            result["ticker"] = clean_ticker
            result["asset_name"] = parts[1]
            result["asset_class"] = asset_class
            return result

    # Single ticker (no dash separator)
    ticker_candidate = parts[0].upper().strip()
    if re.match(r"^[A-Z]{3,5}\d{1,2}F?$", ticker_candidate):
        asset_class, clean_ticker = _classify_ticker(ticker_candidate, "")
        result["ticker"] = clean_ticker
        result["asset_name"] = product
        result["asset_class"] = asset_class
        return result

    return result


def _categorize_row(direction: str, movement_type: str, product: str) -> tuple[str, str]:
    """Return (category, import_as) for a row."""
    mt_norm = _strip_accents(movement_type.lower().strip())

    # Proventos
    for key, import_as in PROVENTO_MAP.items():
        if mt_norm == key or mt_norm.startswith(key):
            return "provento", import_as

    # Renda fixa compra (Credito = incoming asset)
    if mt_norm in RF_COMPRA_TYPES and _strip_accents(direction.lower().strip()) == "credito":
        return "renda_fixa", "compra_rf"

    if mt_norm == "aplicacao":
        return "renda_fixa", "compra_rf"

    if mt_norm == "vencimento":
        return "renda_fixa", "vencimento_rf"

    if mt_norm in ("resgate", "resgate antecipado"):
        return "renda_fixa", "resgate_rf"

    if mt_norm in ("pagamento de juros",):
        return "renda_fixa", "juros_rf"

    if mt_norm in ("amortizacao",):
        return "renda_fixa", "amortizacao_rf"

    # Corporate events
    if "bonificacao" in mt_norm:
        return "evento", "bonificacao"

    if "desdobro" in mt_norm:
        return "evento", "desdobramento"

    # Venda (could be Tesouro or leftover)
    if mt_norm == "venda":
        return "evento", "venda"

    return "ignorado", "ignorado"


# ---------------------------------------------------------------------------
# Preview endpoint
# ---------------------------------------------------------------------------


@router.post("/preview", response_model=MovPreviewResponse)
async def preview_b3_mov(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser .xlsx")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)

    # Try to find Movimentacao sheet
    ws = None
    for name in wb.sheetnames:
        if "movimenta" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows_data: list[MovRow] = []
    header_skipped = False

    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue

        # Skip empty rows
        if not row or not row[0]:
            continue

        # Columns: Entrada/Saida, Data, Movimentacao, Produto, Instituicao, Quantidade, Preco unitario, Valor da Operacao
        if len(row) < 8:
            continue

        direction_val, date_val, mov_type_val, product_val, institution_val, qty_val, price_val, total_val = row[:8]

        direction = str(direction_val or "").strip()
        movement_type = str(mov_type_val or "").strip()
        product = str(product_val or "").strip()
        institution = _abbreviate_broker(str(institution_val or ""))

        try:
            date = _parse_date(date_val)
        except (ValueError, Exception):
            continue

        qty = _parse_number(qty_val)
        unit_price = _parse_number(price_val)
        total_value = _parse_number(total_val)

        # Check if should skip
        skip_reason = _should_skip(movement_type, product)
        if skip_reason:
            rows_data.append(MovRow(
                date=date.isoformat(),
                direction=direction,
                movement_type=movement_type,
                product=product,
                institution=institution,
                qty=qty,
                unit_price=unit_price,
                total_value=total_value,
                category="ignorado",
                import_as="ignorado",
                asset_name=product,
                is_skipped=True,
                skip_reason=skip_reason,
            ))
            continue

        # Categorize
        category, import_as = _categorize_row(direction, movement_type, product)

        # Extract ticker/asset info
        info = _extract_ticker_from_product(product)

        # For renda fixa operations, ensure asset_class is fixed_income
        if category == "renda_fixa" and not info["rf_type"]:
            # Try to detect from product
            info["asset_class"] = "fixed_income"

        # For proventos, if we have a ticker, look up asset class
        # For events, same

        rows_data.append(MovRow(
            date=date.isoformat(),
            direction=direction,
            movement_type=movement_type,
            product=product,
            institution=institution,
            qty=qty,
            unit_price=unit_price,
            total_value=total_value,
            category=category,
            import_as=import_as,
            ticker=info["ticker"],
            asset_name=info["asset_name"],
            asset_class=info["asset_class"],
            rf_type=info["rf_type"],
            rf_code=info["rf_code"],
            is_skipped=(category == "ignorado"),
            skip_reason="Tipo nao suportado" if category == "ignorado" else None,
        ))

    wb.close()

    # --- Duplicate detection ---

    # Load existing dividends for provento duplicate check
    div_result = await db.execute(select(Dividend))
    existing_divs = div_result.scalars().all()
    div_keys = set()
    for d in existing_divs:
        div_keys.add((
            d.date.isoformat() if d.date else "",
            (d.ticker or "").upper(),
            (d.type or "").lower(),
            round(d.value or 0, 2),
        ))

    # Load existing transactions for tx duplicate check
    tx_result = await db.execute(select(Transaction))
    existing_txs = tx_result.scalars().all()
    tx_keys = set()
    for tx in existing_txs:
        tx_keys.add((
            tx.date.isoformat() if tx.date else "",
            (tx.ticker or tx.asset_id or "").upper(),
            tx.operation_type,
            round(tx.qty or 0, 2),
            round(tx.total_value or tx.unit_price or 0, 2),
        ))

    # Check which assets exist
    br_result = await db.execute(select(BrStock.ticker))
    br_tickers = {r[0] for r in br_result.all()}
    fii_result = await db.execute(select(Fii.ticker))
    fii_tickers = {r[0] for r in fii_result.all()}
    etf_result = await db.execute(select(FiEtf.ticker))
    etf_tickers = {r[0] for r in etf_result.all()}
    fi_result = await db.execute(select(FixedIncome.id))
    fi_ids = {r[0] for r in fi_result.all()}

    new_assets = set()

    for row in rows_data:
        if row.is_skipped:
            continue

        # Duplicate check
        if row.category == "provento" or row.import_as in ("juros_rf", "amortizacao_rf"):
            div_type_map = {
                "dividendo": "dividendo",
                "jcp": "jcp",
                "rendimento": "rendimento",
                "juros_rf": "juros",
                "amortizacao_rf": "amortizacao",
            }
            check_ticker = (row.ticker or row.rf_code or "").upper()
            check_type = div_type_map.get(row.import_as, row.import_as)
            key = (row.date, check_ticker, check_type, round(row.total_value or 0, 2))
            if key in div_keys:
                row.is_duplicate = True

        elif row.import_as in ("compra_rf", "vencimento_rf", "resgate_rf"):
            op_map = {"compra_rf": "aporte", "vencimento_rf": "resgate", "resgate_rf": "resgate"}
            check_id = (row.rf_code or "").upper()
            check_op = op_map[row.import_as]
            key = (row.date, check_id, check_op, round(row.qty or 0, 2), round(row.total_value or 0, 2))
            if key in tx_keys:
                row.is_duplicate = True

        elif row.import_as in ("bonificacao", "desdobramento", "venda"):
            check_ticker = (row.ticker or "").upper()
            key = (row.date, check_ticker, row.import_as, round(row.qty or 0, 2), round(row.total_value or row.unit_price or 0, 2))
            if key in tx_keys:
                row.is_duplicate = True

        # Track new assets
        if row.ticker and row.asset_class in ("br_stock", "fii", "fi_etf"):
            asset_sets = {"br_stock": br_tickers, "fii": fii_tickers, "fi_etf": etf_tickers}
            if row.ticker not in asset_sets.get(row.asset_class, set()):
                new_assets.add(row.ticker)
        elif row.rf_code and row.asset_class == "fixed_income":
            if row.rf_code not in fi_ids:
                new_assets.add(f"{row.rf_type or 'RF'}: {row.rf_code}")

    # Build summary
    active = [r for r in rows_data if not r.is_skipped]
    summary = MovSummary(
        total=len(rows_data),
        proventos=sum(1 for r in active if r.category == "provento"),
        renda_fixa=sum(1 for r in active if r.category == "renda_fixa"),
        eventos=sum(1 for r in active if r.category == "evento"),
        ignorados=sum(1 for r in rows_data if r.is_skipped),
        duplicates=sum(1 for r in active if r.is_duplicate),
        new_assets=sorted(new_assets),
    )

    return MovPreviewResponse(rows=rows_data, summary=summary)


# ---------------------------------------------------------------------------
# Confirm endpoint
# ---------------------------------------------------------------------------


@router.post("/confirm", response_model=MovConfirmResponse)
async def confirm_b3_mov(
    body: MovConfirmRequest,
    db: AsyncSession = Depends(get_db),
):
    sorted_rows = sorted(body.rows, key=lambda r: r.date)

    dividends_created = 0
    transactions_created = 0
    assets_created = []
    errors = []

    # Pre-load existing assets
    br_result = await db.execute(select(BrStock.ticker))
    br_tickers = {r[0] for r in br_result.all()}
    fii_result = await db.execute(select(Fii.ticker))
    fii_tickers = {r[0] for r in fii_result.all()}
    etf_result = await db.execute(select(FiEtf.ticker))
    etf_tickers = {r[0] for r in etf_result.all()}
    fi_result = await db.execute(select(FixedIncome.id))
    fi_ids = {r[0] for r in fi_result.all()}

    asset_sets = {"br_stock": br_tickers, "fii": fii_tickers, "fi_etf": etf_tickers}

    # Pre-fetch sector info from Yahoo Finance for new tickers
    new_tickers_to_lookup = set()
    for row in sorted_rows:
        if row.ticker and row.asset_class in ("br_stock", "fii"):
            ticker_set = asset_sets.get(row.asset_class, set())
            if row.ticker not in ticker_set:
                new_tickers_to_lookup.add(row.ticker)

    asset_info_map = {}
    if new_tickers_to_lookup:
        try:
            asset_info_map = await fetch_asset_info(list(new_tickers_to_lookup))
        except Exception:
            pass  # Fallback: will use "A classificar"

    for row in sorted_rows:
        try:
            date = datetime.date.fromisoformat(row.date)

            # --- Proventos ---
            if row.import_as in ("dividendo", "jcp", "rendimento"):
                div = Dividend(
                    date=date,
                    ticker=(row.ticker or row.rf_code or row.asset_name)[:10],
                    type=row.import_as,
                    value=row.total_value or 0,
                )
                db.add(div)
                dividends_created += 1

            # --- Renda Fixa compra ---
            elif row.import_as == "compra_rf":
                # Create FixedIncome asset if not exists
                if row.rf_code and row.rf_code not in fi_ids:
                    fi = FixedIncome(
                        id=row.rf_code[:36],
                        title=row.product[:120],
                        type=row.rf_type or "Outro",
                        rate="A definir",
                        applied_value=0,
                        current_value=0,
                        application_date=date,
                        maturity_date=date + datetime.timedelta(days=365 * 3),
                        broker=row.institution[:30],
                        indexer="CDI",
                        contracted_rate=0,
                        tax_exempt=False,
                    )
                    db.add(fi)
                    fi_ids.add(row.rf_code)
                    assets_created.append(f"{row.rf_type or 'RF'}: {row.rf_code}")
                    await db.flush()

                # Create aporte transaction
                tx = Transaction(
                    date=date,
                    operation_type="aporte",
                    asset_class="fixed_income",
                    ticker=None,
                    asset_id=row.rf_code[:36] if row.rf_code else None,
                    asset_name=(row.asset_name or row.product)[:200],
                    qty=None,
                    unit_price=None,
                    total_value=row.total_value or 0,
                    broker=row.institution[:30],
                    fees=0,
                    notes=f"Importado B3 Mov - {row.movement_type}",
                )
                db.add(tx)
                await db.flush()
                await _apply(db, tx)
                transactions_created += 1

            # --- Renda Fixa vencimento / resgate ---
            elif row.import_as in ("vencimento_rf", "resgate_rf"):
                tv = row.total_value or 0
                tx = Transaction(
                    date=date,
                    operation_type="resgate",
                    asset_class="fixed_income",
                    ticker=None,
                    asset_id=row.rf_code[:36] if row.rf_code else None,
                    asset_name=(row.asset_name or row.product)[:200],
                    qty=None,
                    unit_price=None,
                    total_value=tv,
                    broker=row.institution[:30],
                    fees=0,
                    notes=f"Importado B3 Mov - {row.movement_type}",
                )
                db.add(tx)
                await db.flush()

                # For resgate with a known value, set current_value to the
                # actual redemption amount so the return shows correctly
                # (the standard _apply subtracts, which produces wrong results
                # when the rate is unknown and current_value == applied_value).
                asset = await db.get(FixedIncome, row.rf_code[:36]) if row.rf_code else None
                if asset and tv > 0 and asset.contracted_rate == 0:
                    # Redemption value known: set as current_value
                    asset.current_value = tv
                    asset.maturity_date = date
                elif asset and tv == 0 and row.import_as == "vencimento_rf":
                    # Vencimento with no value reported: update maturity date
                    asset.maturity_date = date
                    # Don't call _apply (total=0 wouldn't change anything)
                else:
                    await _apply(db, tx)

                transactions_created += 1

            # --- Renda Fixa juros / amortizacao ---
            elif row.import_as in ("juros_rf", "amortizacao_rf"):
                div_type = "juros" if row.import_as == "juros_rf" else "amortizacao"
                div = Dividend(
                    date=date,
                    ticker=(row.rf_code or row.asset_name)[:10],
                    type=div_type,
                    value=row.total_value or 0,
                )
                db.add(div)
                dividends_created += 1

            # --- Corporate events: bonificacao, desdobramento ---
            elif row.import_as in ("bonificacao", "desdobramento"):
                if row.ticker and row.asset_class:
                    # Auto-create asset if missing
                    ticker_set = asset_sets.get(row.asset_class, set())
                    if row.ticker not in ticker_set:
                        info = asset_info_map.get(row.ticker, {})
                        sector = info.get("sector") or "A classificar"
                        name = info.get("name") or row.asset_name or row.ticker

                        if row.asset_class == "br_stock":
                            db.add(BrStock(
                                ticker=row.ticker, name=name,
                                sector=sector, qty=0, avg_price=0,
                                current_price=0, broker=row.institution[:30],
                            ))
                        elif row.asset_class == "fii":
                            db.add(Fii(
                                ticker=row.ticker, name=name,
                                sector=sector, qty=0, avg_price=0,
                                current_price=0, broker=row.institution[:30],
                            ))
                        elif row.asset_class == "fi_etf":
                            db.add(FiEtf(
                                ticker=row.ticker, name=info.get("name") or row.asset_name or row.ticker,
                                qty=0, avg_price=0, current_price=0,
                                broker=row.institution[:30],
                            ))
                        ticker_set.add(row.ticker)
                        assets_created.append(row.ticker)
                        await db.flush()

                    tx = Transaction(
                        date=date,
                        operation_type=row.import_as,
                        asset_class=row.asset_class,
                        ticker=row.ticker,
                        asset_id=None,
                        asset_name=(row.asset_name or row.ticker)[:200],
                        qty=row.qty or 0,
                        unit_price=row.unit_price,
                        total_value=row.total_value,
                        broker=row.institution[:30],
                        fees=0,
                        notes=f"Importado B3 Mov - {row.movement_type}",
                    )
                    db.add(tx)
                    await db.flush()
                    await _apply(db, tx)
                    transactions_created += 1

            # --- Venda (Tesouro or stocks) ---
            elif row.import_as == "venda":
                if row.rf_code and row.asset_class == "fixed_income":
                    # Treat as resgate
                    tv = row.total_value or 0
                    tx = Transaction(
                        date=date,
                        operation_type="resgate",
                        asset_class="fixed_income",
                        ticker=None,
                        asset_id=row.rf_code[:36] if row.rf_code else None,
                        asset_name=(row.asset_name or row.product)[:200],
                        qty=None,
                        unit_price=None,
                        total_value=tv,
                        broker=row.institution[:30],
                        fees=0,
                        notes=f"Importado B3 Mov - Venda",
                    )
                    db.add(tx)
                    await db.flush()

                    # Same as resgate: set current_value directly for rate-unknown bonds
                    asset = await db.get(FixedIncome, row.rf_code[:36]) if row.rf_code else None
                    if asset and tv > 0 and asset.contracted_rate == 0:
                        asset.current_value = tv
                        asset.maturity_date = date
                    else:
                        await _apply(db, tx)

                    transactions_created += 1
                elif row.ticker and row.asset_class:
                    # Stock venda
                    ticker_set = asset_sets.get(row.asset_class, set())
                    if row.ticker not in ticker_set:
                        info = asset_info_map.get(row.ticker, {})
                        sector = info.get("sector") or "A classificar"
                        name = info.get("name") or row.asset_name or row.ticker

                        if row.asset_class == "br_stock":
                            db.add(BrStock(
                                ticker=row.ticker, name=name,
                                sector=sector, qty=0, avg_price=0,
                                current_price=0, broker=row.institution[:30],
                            ))
                        elif row.asset_class == "fii":
                            db.add(Fii(
                                ticker=row.ticker, name=name,
                                sector=sector, qty=0, avg_price=0,
                                current_price=0, broker=row.institution[:30],
                            ))
                        elif row.asset_class == "fi_etf":
                            db.add(FiEtf(
                                ticker=row.ticker, name=info.get("name") or row.asset_name or row.ticker,
                                qty=0, avg_price=0, current_price=0,
                                broker=row.institution[:30],
                            ))
                        ticker_set.add(row.ticker)
                        assets_created.append(row.ticker)
                        await db.flush()

                    tx = Transaction(
                        date=date,
                        operation_type="venda",
                        asset_class=row.asset_class,
                        ticker=row.ticker,
                        asset_id=None,
                        asset_name=(row.asset_name or row.ticker)[:200],
                        qty=row.qty or 0,
                        unit_price=row.unit_price or 0,
                        total_value=row.total_value or 0,
                        broker=row.institution[:30],
                        fees=0,
                        notes=f"Importado B3 Mov - Venda",
                    )
                    db.add(tx)
                    await db.flush()
                    await _apply(db, tx)
                    transactions_created += 1

        except Exception as e:
            errors.append(f"{row.product} ({row.date}): {str(e)}")

    await db.commit()

    return MovConfirmResponse(
        dividends_created=dividends_created,
        transactions_created=transactions_created,
        assets_created=sorted(set(assets_created)),
        errors=errors,
    )
